import cv2
import numpy as np
import math
import argparse
import os
from openpyxl import Workbook
import sys
import time

from openpyxl.styles import PatternFill

import Histo
import Matches


def progressbar(it, prefix="", size=60, file=sys.stdout):
    count = len(it)

    def show(j):
        x = int(size * j / count)
        file.write("%s[%s%s] %i/%i\r" % (prefix, "#" * x, "." * (size - x), j, count))
        file.flush()

    show(0)
    for i, item in enumerate(it):
        yield item
        show(i + 1)
    file.write("\n")
    file.flush()


# Fonction permettant de trouver les logos dans une image
def Draw(Images, name, dataset):
    img_lst = list()  # Liste des images ouvertes
    img_color = list()
    orb_img = cv2.ORB_create(nfeatures=10000, edgeThreshold=1)
    bf = cv2.BFMatcher(cv2.NORM_HAMMING)

    outputImg = []

    if Images is not None:
        for image in Images:
            img = cv2.imread(image, cv2.IMREAD_COLOR)
            img_lst.append(img)
            img_color.append(list(np.random.choice(range(256), size=3)))

    des_list = list()
    kp_lst = list()
    for image in img_lst:
        kp, des = orb_img.detectAndCompute(image, None)
        des_list.append(des)
        kp_lst.append(kp)
    for img in progressbar(dataset, "Analysing images ... : ", 60):
        img_dataset = cv2.imread(img, cv2.IMREAD_COLOR)
        if (img_dataset.shape[0] < 1080 and img_dataset.shape[1] < 1920) or (
                img_dataset.shape[0] > 1500 and img_dataset.shape[1] > 2500):
            img_dataset = cv2.resize(img_dataset, (1920, 1080), interpolation=cv2.INTER_CUBIC)
        cleanCopy = img_dataset
        kp2, des2 = orb_img.detectAndCompute(img_dataset, None)
        # Partie Homography
        # Itération sur chaque image
        bestMatche = list()
        bestCor = 0
        bestLogo = ''
        # print("______________________________________")
        for index_kp in range(len(kp_lst)):
            goodpoints = Matches.GoodPoints(des_list[index_kp], des2, bf)
            cor = Histo.matchRateMatches(goodpoints)
            if cor > bestCor:
                bestCor = cor
                bestLogo = name[index_kp]
            # print("cor: " + str(cor) + "; name : " + name[index_kp])
            # print("Nb good : " + str(len(goodpoints)) + "; name : " + name[index_kp])
            if len(goodpoints) > 25:
                sch_pts = np.float32([kp_lst[index_kp][m.queryIdx].pt for m in goodpoints]).reshape(-1, 1, 2)
                img_pts = np.float32([kp2[m.trainIdx].pt for m in goodpoints]).reshape(-1, 1, 2)
                matrix, mask = cv2.findHomography(sch_pts, img_pts, cv2.RANSAC, 5.0)

                if matrix is not None:
                    pts = np.float32([[0, 0],
                                      [0, img_lst[index_kp].shape[0] - 1],
                                      [img_lst[index_kp].shape[1] - 1, img_lst[index_kp].shape[0] - 1],
                                      [img_lst[index_kp].shape[1] - 1, 0]]).reshape(-1, 1, 2)
                    dst = cv2.perspectiveTransform(pts, matrix)
                    rect = cv2.minAreaRect(dst)

                    box = cv2.boxPoints(rect)
                    box = np.int0(box)

                    col = img_color[index_kp]
                    # Filtre par angle pour prendre des régions cohérentes
                    angle_bas = Matches.getAngle(dst[0][0], dst[1][0], dst[2][0])
                    angle_haut = Matches.getAngle(pts[2][0], pts[3][0], pts[0][0])
                    if (70 < angle_bas < 110) and (70 < angle_haut < 110):
                        xmin = min([pt[0] for pt in box])
                        xmax = max([pt[0] for pt in box])
                        ymin = min([pt[1] for pt in box])
                        ymax = max([pt[1] for pt in box])
                        boundingBox = (ymin, ymax, xmin, xmax)
                        # Calcul de l'histogramme de l'image trouvée et du logo à comparer
                        himg = Histo.histo(cleanCopy, boundingBox)
                        imgResized = cv2.resize(img_lst[index_kp], (int(ymax - ymin), int(xmax - xmin)))
                        hLogo = Histo.histo(imgResized)

                        # Application de flou gaussien sur l'histogramme
                        himg = Histo.gaussianblur(himg)
                        hLogo = Histo.gaussianblur(hLogo)
                        # Passage de l'histogramme en 8x256
                        himg = Histo.cutHisto(himg)
                        hLogo = Histo.cutHisto(hLogo)

                        # Récupération du pourcentage de différence entre deux histogrammes
                        r = Histo.matchRateHisto(himg, hLogo, imgResized.shape[0] * imgResized.shape[1])
                        bestMatche.append((r, index_kp, col, box))
                        # print(str((1 - r) * 100) + " de difference")
        bottomLeftCornerOfTextCompa = (int(50), int(50))
        bottomLeftCornerOfTextColor = (int(50), int(100))
        bottomLeftCornerOfTextLogo = (int(50), int(150))

        fontScale = 2
        fontColor = (0, 255, 0)
        thickness = 3
        lineType = 2
        nameLogo = ''
        if len(bestMatche) > 0:
            best = min([tuple[0] for tuple in bestMatche])
            bestTuple = [tuple for tuple in bestMatche if tuple[0] == best]
            nameLogo = name[bestTuple[0][1]]
            x_offset = 50
            y_offset = 200
            logoResized = cv2.resize(img_lst[bestTuple[0][1]],
                                     (int(img_dataset.shape[1] / 10), int(img_dataset.shape[0] / 10)))
            img_dataset[y_offset:y_offset + logoResized.shape[0],
            x_offset:x_offset + logoResized.shape[1]] = logoResized
            cv2.drawContours(img_dataset, [bestTuple[0][3]], 0,
                             [int(bestTuple[0][2][0]), int(bestTuple[0][2][1]), int(bestTuple[0][2][2])], 2)
            img_dataset = cv2.putText(img_dataset,
                                      "Pourcentage compatibilite Color : " + str(bestTuple[0][0][0] * 100) + " %",
                                      bottomLeftCornerOfTextColor,
                                      1,
                                      fontScale,
                                      fontColor,
                                      thickness,
                                      lineType)

        img_dataset = cv2.putText(img_dataset, "Pourcentage compatibilite Matches : " + str(bestCor * 100) + " %",
                                  bottomLeftCornerOfTextCompa,
                                  1,
                                  fontScale,
                                  fontColor,
                                  thickness,
                                  lineType)

        img_dataset = cv2.putText(img_dataset, "Logo trouve : " + bestLogo,
                                  bottomLeftCornerOfTextLogo,
                                  1,
                                  fontScale,
                                  fontColor,
                                  thickness,
                                  lineType)
        outputLogo = ''
        if nameLogo != '':
            outputLogo = nameLogo
        else:
            outputLogo = bestLogo
        outputImg.append((img_dataset, outputLogo, str(bestCor * 100), str(bestTuple[0][0][0] * 100)))
    return outputImg


# Fonction permettant de récuperer les logos ainsi que le dataset
def getImages(Logos, DataSet):
    Images = list()
    Images_DataSet = list()
    name = list()
    nameDataset = list()
    # DataSet = "Logo/Dessins/"
    for (root, sousRep, fic) in os.walk(Logos):
        for dirname in sousRep:
            for (repertoire, sousRepertoires, fichiers) in os.walk(os.path.join(root, dirname)):
                for nameFile in fichiers:
                    if nameFile.split('.')[1] == "jpg" or nameFile.split('.')[1] == "png":
                        pathImg = repertoire + '/' + nameFile
                        name.append(dirname)
                        Images.append(pathImg)

    for (root, sousRep, fic) in os.walk(DataSet):
        for dirname in sousRep:
            for (repertoire, sousRepertoires, fichiers) in os.walk(os.path.join(root, dirname)):
                for nameFile in fichiers:
                    if nameFile.split('.')[1] == "jpg" or nameFile.split('.')[1] == "png":
                        pathImg = repertoire + '/' + nameFile
                        nameDataset.append(dirname)
                        Images_DataSet.append(pathImg)
    return Images, name, Images_DataSet, nameDataset


def haarCascade(Images_DataSet):
    a = time.time()
    output = list()
    for image in Images_DataSet:
        img = cv2.imread(image, cv2.IMREAD_COLOR)
        coca_cascade = cv2.CascadeClassifier('cascade_clasifier/trained_classifier/coca_cascade_15_train.xml')
        adidas_cascade = cv2.CascadeClassifier('cascade_clasifier/trained_classifier/adadas_cascade_15_train.xml')
        nike_cascade = cv2.CascadeClassifier('cascade_clasifier/trained_classifier/nike_cascade_15_train.xml')
        if (img.shape[0] < 1080 and img.shape[1] < 1920) or (
                img.shape[0] > 1500 and img.shape[1] > 2500):
            img = cv2.resize(img, (1920, 1080), interpolation=cv2.INTER_CUBIC)

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        coca = coca_cascade.detectMultiScale(gray,
                                             scaleFactor=1.1,
                                             minNeighbors=5,
                                             minSize=(30, 30),  # (60, 20),
                                             flags=cv2.CASCADE_SCALE_IMAGE)
        adidas = adidas_cascade.detectMultiScale(gray,
                                                 scaleFactor=1.1,
                                                 minNeighbors=5,
                                                 minSize=(30, 30),  # (60, 20),
                                                 flags=cv2.CASCADE_SCALE_IMAGE)
        nike = nike_cascade.detectMultiScale(gray,
                                             scaleFactor=1.1,
                                             minNeighbors=5,
                                             minSize=(30, 30),  # (60, 20),
                                             flags=cv2.CASCADE_SCALE_IMAGE)
        outputLogo = ''
        sizeCoca = len(coca)
        sizeAdidas = len(adidas)
        sizeNike = len(nike)
        sizeArray = [(sizeCoca, 'Coca-Cola'), (sizeAdidas, 'Adidas'), (sizeNike, 'Nike')]
        maxSize = max([sizeCoca, sizeAdidas, sizeNike])
        best = [tuple for tuple in sizeArray if tuple[0] == maxSize]
        outputLogo = best[0][1]
        for (x, y, w, h) in coca:
            cv2.rectangle(img, (x, y), (x + w, y + h), (0, 0, 255), 2)
        for (x, y, w, h) in adidas:
            cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
        for (x, y, w, h) in nike:
            cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)
        output.append((img, outputLogo))
    return output


def writeExcel(output, output2, logoExpected, execution_time1, execution_time2):
    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    greenFill = PatternFill(start_color='FF8FCE00',
                            end_color='FF8FCE00',
                            fill_type='solid')
    blueFill = PatternFill(start_color='FF3852A4',
                            end_color='FF3852A4',
                            fill_type='solid')

    columns = ["A", "B", "C", "D"]
    # Create a Workbook on Excel:
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'data'
    sheet.column_dimensions["A"].width = 13
    sheet.column_dimensions["B"].width = 19
    sheet.column_dimensions["C"].width = 19
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 20
    sheet.column_dimensions["F"].width = 21
    sheet.column_dimensions["G"].width = 21
    sheet.column_dimensions["I"].width = 20
    sheet.column_dimensions["J"].width = 20
    sheet.column_dimensions["K"].width = 25
    sheet.column_dimensions["L"].width = 28


    # Print the titles into Excel Workbook:
    row = 1
    sheet['A' + str(row)] = 'Logo Expected'
    sheet['A' + str(row)].fill = blueFill
    sheet['B' + str(row)] = 'Logo Found Method 1'
    sheet['B' + str(row)].fill = blueFill
    sheet['C' + str(row)] = 'Logo Found Method 2'
    sheet['C' + str(row)].fill = blueFill
    sheet['D' + str(row)] = 'Matching Rate Color'
    sheet['D' + str(row)].fill = blueFill
    sheet['E' + str(row)] = 'Matching Rate Matches'
    sheet['E' + str(row)].fill = blueFill
    sheet['F' + str(row)] = 'False-Positive Method 1'
    sheet['F' + str(row)].fill = blueFill
    sheet['G' + str(row)] = 'False-Positive Method 2'
    sheet['G' + str(row)].fill = blueFill

    sheet['K' + str(row)] = "Execution time First method"
    sheet['K' + str(row)].fill = blueFill
    sheet['K' + str(row + 1)] = str(execution_time1) + " sec"

    sheet['L' + str(row)] = "Execution time Second method"
    sheet['L' + str(row)].fill = blueFill
    sheet['L' + str(row + 1)] = str(execution_time2) + " sec"


    goodAnswer1 = 0
    goodAnswer2 = 0
    # Populate with data
    for i in range(len(output)):
        row += 1
        sheet['A' + str(row)] = logoExpected[i]
        sheet['B' + str(row)] = output[i][1]
        sheet['C' + str(row)] = output2[i][1]
        sheet['D' + str(row)] = output[i][2] + " %"
        sheet['E' + str(row)] = output[i][3] + " %"
        if logoExpected[i] != output[i][1]:
            sheet['F' + str(row)] = "True"
            sheet['F' + str(row)].fill = redFill
        else:
            sheet['F' + str(row)] = "False"
            sheet['F' + str(row)].fill = greenFill
            goodAnswer1 += 1
        if logoExpected[i] != output2[i][1]:
            sheet['G' + str(row)] = "True"
            sheet['G' + str(row)].fill = redFill
        else:
            sheet['G' + str(row)] = "False"
            sheet['G' + str(row)].fill = greenFill
            goodAnswer2 += 1

    sheet['I' + str(1)] = "Good Answers rate M1"
    sheet['I' + str(1)].fill = blueFill
    sheet['I' + str(2)] = str((goodAnswer1 / len(output)) * 100) + " %"

    sheet['J' + str(1)] = "Good Answers rate M2"
    sheet['J' + str(1)].fill = blueFill
    sheet['J' + str(2)] = str((goodAnswer2 / len(output)) * 100) + " %"
    filename = 'data_ImgLogo.xlsx'
    wb.save(filename)

    # Open the file for the user:
    os.chdir(sys.path[0])
    os.system('start excel.exe "%s\\%s"' % (sys.path[0], filename,))


def main():
    a = time.time()
    Logos = "Logo/Clair/"
    DataSet = "Logo/SurPhotos/"
    Images, name, Images_DataSet, nameDataset = getImages(Logos, DataSet)
    output = Draw(Images, name, Images_DataSet)

    execution_time = format(time.time() - a, '.2f')
    print("Execution time First method: " + str(execution_time) + "sec")
    b = time.time()
    output2 = haarCascade(Images_DataSet)
    execution_time2 = format(time.time() - b, '.2f')
    print("Execution time Second method: " + str(execution_time2) + "sec")
    time.sleep(3)
    writeExcel(output, output2, nameDataset, execution_time, execution_time2)
    for i in range(len(output)):
        cv2.imshow("results :", output[i][0])
        cv2.imshow("results 2:", output2[i][0])
        cv2.waitKey(0)


if __name__ == "__main__":
    main()
